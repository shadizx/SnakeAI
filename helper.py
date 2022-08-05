import matplotlib.pyplot as plt
from IPython import display
from openpyxl import load_workbook

plt.ion()

def plot(scores, mean_scores):
    display.clear_output(wait=True)
    display.display(plt.gcf())
    plt.clf()
    plt.title('Training...')
    plt.xlabel('Number of games')
    plt.ylabel('Score')
    plt.plot(scores)
    plt.plot(mean_scores)
    plt.ylim(ymin=0)
    plt.text(len(scores) - 1, scores[-1], str(scores[-1]))
    plt.text(len(mean_scores) - 1, mean_scores[-1], str(mean_scores[-1]))

def annotate(stats, worksheet):
    book = load_workbook(worksheet)
    sheet = book.active
    row = sheet.max_row+1
    sheet[row][0].value = row-1  # fixes attempt number
    for index, cell in enumerate(sheet[row][1:14]):
        cell.value = stats[index]
    book.save(worksheet)